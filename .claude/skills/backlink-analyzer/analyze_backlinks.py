#!/usr/bin/env python3
"""
backlink-analyzer / analyze_backlinks.py

Phân tích file Ahrefs "Referring Domains" export (1 hoặc nhiều site).
- Tự dò mã hóa (UTF-16 / UTF-8-SIG) + delimiter (tab/phẩy).
- Map cột linh hoạt theo alias (xem config/ahref-columns.md).
- Xuất báo cáo Excel nhiều sheet + JSON tóm tắt để Agent viết báo cáo Markdown.

Cần openpyxl (giống skill keyword-expand). Nếu thiếu: pip3 install openpyxl.

Cách dùng:
    python3 analyze_backlinks.py <file1.csv> [file2.csv ...] \
        [--mine <file_cua_minh.csv>] --out outputs/backlink.xlsx --json outputs/backlink.json
"""
import sys, os, csv, io, json, argparse, statistics, re
from datetime import datetime, timezone

# ------- alias cột (khớp với config/ahref-columns.md) -------
ALIASES = {
    "domain": ["domain", "referring domain", "referring domains"],
    "is_spam": ["is spam", "spam"],
    "dr": ["dr", "domain rating"],
    "ref_domains": ["dofollow ref. domains", "referring domains"],
    "linked_domains": ["dofollow linked domains", "linked domains"],
    "traffic": ["traffic", "domain traffic"],
    "keywords": ["keywords", "organic keywords"],
    "links_to_target": ["links to target"],
    "dofollow_links": ["dofollow links"],
    "first_seen": ["first seen", "first seen link"],
    "lost": ["lost", "last seen"],
}


def read_rows(path):
    """Đọc file, tự dò encoding + delimiter. Trả (header_list, list[dict])."""
    raw = open(path, "rb").read()
    text = None
    for enc in ("utf-16", "utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        raise ValueError(f"Không decode được file {path}")
    first = text.splitlines()[0] if text.splitlines() else ""
    delim = "\t" if first.count("\t") >= first.count(",") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        return [], []
    header = [h.strip() for h in rows[0]]
    dicts = [dict(zip(header, r)) for r in rows[1:] if r and any(c.strip() for c in r)]
    return header, dicts


def build_colmap(header):
    """Map trường chuẩn -> tên cột thực trong file."""
    lower = {h.lower().strip(): h for h in header}
    colmap = {}
    for field, al in ALIASES.items():
        for a in al:
            if a in lower:
                colmap[field] = lower[a]
                break
    return colmap


def num(x):
    if x is None:
        return None
    x = str(x).strip().replace(",", "")
    if x == "":
        return None
    try:
        return float(x)
    except ValueError:
        return None


def get(row, colmap, field):
    col = colmap.get(field)
    return row.get(col, "") if col else ""


def is_toxic(row, colmap):
    reasons = []
    spam = str(get(row, colmap, "is_spam")).strip().lower()
    if spam == "true":
        reasons.append("Ahrefs gắn cờ spam")
    dr = num(get(row, colmap, "dr"))
    tr = num(get(row, colmap, "traffic"))
    ld = num(get(row, colmap, "linked_domains"))
    if dr is not None and dr < 10 and (tr == 0 or tr is None):
        reasons.append("DR<10 & traffic=0")
    if ld is not None and ld > 5000 and dr is not None and dr < 20:
        reasons.append("nghi link farm/PBN (linked_domains>5000, DR<20)")
    return reasons


def analyze_file(path):
    header, rows = read_rows(path)
    colmap = build_colmap(header)
    missing = [f for f in ("domain", "dr") if f not in colmap]
    site = os.path.basename(path).split("-refdomains")[0].split("_")[0]

    total = len(rows)
    drs, toxic, dofollow, live, lost = [], 0, 0, 0, 0
    dr_buckets = {"strong(70-100)": 0, "good(30-69)": 0, "low(10-29)": 0, "verylow(0-9)": 0}
    enriched = []
    for r in rows:
        dr = num(get(r, colmap, "dr"))
        tr = num(get(r, colmap, "traffic"))
        dfl = num(get(r, colmap, "dofollow_links"))
        lost_v = str(get(r, colmap, "lost")).strip()
        reasons = is_toxic(r, colmap)
        if dr is not None:
            drs.append(dr)
            if dr >= 70: dr_buckets["strong(70-100)"] += 1
            elif dr >= 30: dr_buckets["good(30-69)"] += 1
            elif dr >= 10: dr_buckets["low(10-29)"] += 1
            else: dr_buckets["verylow(0-9)"] += 1
        if reasons: toxic += 1
        if dfl and dfl > 0: dofollow += 1
        if lost_v: lost += 1
        else: live += 1
        enriched.append({
            "domain": get(r, colmap, "domain"),
            "dr": dr, "traffic": tr,
            "dofollow": bool(dfl and dfl > 0),
            "toxic": bool(reasons),
            "toxic_reason": "; ".join(reasons),
            "first_seen": get(r, colmap, "first_seen"),
            "lost": lost_v,
        })

    summary = {
        "site": site,
        "file": os.path.basename(path),
        "columns_found": colmap,
        "columns_missing": missing,
        "total_ref_domains": total,
        "toxic_count": toxic,
        "toxic_pct": round(toxic / total * 100, 1) if total else 0,
        "dofollow_count": dofollow,
        "dofollow_pct": round(dofollow / total * 100, 1) if total else 0,
        "live": live, "lost": lost,
        "dr_median": round(statistics.median(drs), 1) if drs else None,
        "dr_max": max(drs) if drs else None,
        "dr_buckets": dr_buckets,
    }
    top = sorted([e for e in enriched if e["dr"] is not None],
                 key=lambda e: (e["dr"], e["traffic"] or 0), reverse=True)[:20]
    top_toxic = [e for e in enriched if e["toxic"]]
    top_toxic = sorted(top_toxic, key=lambda e: (e["traffic"] or 0), reverse=True)[:20]
    summary["top_domains"] = top
    summary["top_toxic"] = top_toxic
    return summary, enriched


def write_xlsx(analyses, gap, out):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[!] Thiếu openpyxl. Chạy: pip3 install openpyxl", file=sys.stderr)
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    hdr = Font(bold=True)
    ws.append(["Site", "Total", "Toxic", "Toxic %", "Dofollow", "Dofollow %",
               "Live", "Lost", "DR median", "DR max"])
    for c in ws[1]: c.font = hdr
    for a in analyses:
        ws.append([a["site"], a["total_ref_domains"], a["toxic_count"], a["toxic_pct"],
                   a["dofollow_count"], a["dofollow_pct"], a["live"], a["lost"],
                   a["dr_median"], a["dr_max"]])
    # per-site sheet
    for a in analyses:
        name = re.sub(r"[^A-Za-z0-9]", "_", a["site"])[:28] or "site"
        s = wb.create_sheet(name)
        s.append(["Domain", "DR", "Traffic", "Dofollow", "Toxic", "Reason", "First seen", "Lost"])
        for c in s[1]: c.font = hdr
        for e in a["_all"]:
            s.append([e["domain"], e["dr"], e["traffic"], "yes" if e["dofollow"] else "no",
                      "yes" if e["toxic"] else "no", e["toxic_reason"], e["first_seen"], e["lost"]])
        s.freeze_panes = "A2"
        s.auto_filter.ref = f"A1:H{s.max_row}"
    if gap is not None:
        g = wb.create_sheet("Link Gap")
        g.append(["Domain", "DR", "Traffic", "Trỏ về (đối thủ)"])
        for c in g[1]: c.font = hdr
        for e in gap:
            g.append([e["domain"], e["dr"], e["traffic"], e["from"]])
        g.freeze_panes = "A2"
    wb.save(out)
    return True


def compute_gap(analyses, mine_site):
    """Domain trỏ về đối thủ nhưng không có ở site của mình, bỏ nghi độc."""
    mine = next((a for a in analyses if a["site"] == mine_site), None)
    if not mine or len(analyses) < 2:
        return None
    mine_domains = {e["domain"].lower() for e in mine["_all"]}
    gap = {}
    for a in analyses:
        if a["site"] == mine_site:
            continue
        for e in a["_all"]:
            d = e["domain"].lower()
            if d in mine_domains or e["toxic"]:
                continue
            cur = gap.get(d)
            if not cur or (e["dr"] or 0) > (cur["dr"] or 0):
                gap[d] = {"domain": e["domain"], "dr": e["dr"], "traffic": e["traffic"], "from": a["site"]}
    return sorted(gap.values(), key=lambda e: (e["dr"] or 0, e["traffic"] or 0), reverse=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("--mine", help="file của site mình (để tính link gap)")
    ap.add_argument("--out", default="backlink.xlsx")
    ap.add_argument("--json", help="ghi JSON tóm tắt")
    args = ap.parse_args()

    analyses = []
    for path in args.files:
        if not os.path.exists(path):
            print(f"[!] Không thấy file: {path}", file=sys.stderr)
            continue
        summ, enriched = analyze_file(path)
        summ["_all"] = enriched
        analyses.append(summ)
    if not analyses:
        print("[!] Không có file hợp lệ.", file=sys.stderr)
        sys.exit(1)

    mine_site = None
    if args.mine:
        mine_site = os.path.basename(args.mine).split("-refdomains")[0].split("_")[0]
    gap = compute_gap(analyses, mine_site) if mine_site else None

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    ok = write_xlsx(analyses, gap, args.out)

    # JSON tóm tắt (bỏ _all cho gọn, giữ top_*)
    slim = []
    for a in analyses:
        c = {k: v for k, v in a.items() if k != "_all"}
        slim.append(c)
    payload = {"analyses": slim, "link_gap_top": (gap or [])[:30] if gap else None}
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2)

    for a in slim:
        print(f"[✓] {a['site']}: {a['total_ref_domains']} ref domains | "
              f"nghi độc {a['toxic_count']} ({a['toxic_pct']}%) | "
              f"dofollow {a['dofollow_pct']}% | DR median {a['dr_median']}")
        if a["columns_missing"]:
            print(f"    [!] Thiếu cột: {a['columns_missing']}")
    if gap is not None:
        print(f"[✓] Link gap: {len(gap)} cơ hội (đã lọc nghi độc)")
    if ok:
        print(f"[✓] Excel: {args.out}")
    if args.json:
        print(f"[✓] JSON: {args.json}")


if __name__ == "__main__":
    main()
